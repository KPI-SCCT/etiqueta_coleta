import re
from datetime import datetime
from io import BytesIO
from pathlib import Path

import streamlit as st
from etiqueta_layout_engine import draw_template_padrao, draw_template_rede

try:
    from reportlab.graphics.barcode import code128
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    REPORTLAB_AVAILABLE = True
except Exception:
    code128 = None
    A4 = None
    canvas = None
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except Exception:
    load_workbook = None
    OPENPYXL_AVAILABLE = False


APP_NAME = "COLETA"
PROJETO_REDE = "REDE"
MIDIA_A4 = "A4 - varias etiquetas por folha"
MIDIA_TERMICA = "Etiqueta termica 100x80 - 1 por pagina"
PLANILHA_BASE_CRED = "bases padrão + cred.xlsx"

DESTINOS = [
    "CTDI DO BR - SP",
    "FLEXTRONIC",
    "FEDEX CAJAMAR - SP",
    "DHL LOUVEIRA - SP",
]

PROJETOS = [
    "CIELO - POS",
    "CIELO - TEF",
    "CIELO - TRANSF",
    "FISERV",
    "MOOZ",
    "STONE",
    "PICPAY",
    "PAGBANK",
    "CTRENDS",
    "C6BANK",
    "ADYEN",
    "CLOUDWALK",
    PROJETO_REDE,
]

PREFIXOS_ROMANEIO = {
    "CIELO - POS": "1.2/",
    "CIELO - TEF": "2.2/",
    "CIELO - TRANSF": "1.3/",
    "FISERV": "34.3/",
    "MOOZ": "42.3/",
    "STONE": "41.3/",
    "PICPAY": "49.3/",
    "PAGBANK": "53.3/",
    "CTRENDS": "39.3/",
    "C6BANK": "43.3/",
    "ADYEN": "45.3/",
    "CLOUDWALK": "40.3/",
    PROJETO_REDE: "51.2/",
}

CRED_OPTIONS = [
    ("CRED369", "POLO REDE PONTA GROSSA"),
    ("CRED385", "POLO REDE SJ DOS CAMPOS"),
    ("CRED368", "POLO REDE SANTOS"),
    ("CRED372", "POLO REDE CUR PINHAIS"),
    ("CRED382", "POLO REDE LONDRINA"),
    ("CRED371", "POLO REDE CURITIBA"),
    ("CRED370", "POLO REDE MARINGA"),
    ("CRED384", "POLO REDE PINDA"),
    ("CRED383", "POLO REDE CASCAVEL"),
    ("CRED408", "POLO REDE FORTALEZA"),
    ("CRED409", "POLO REDE JUAZ DO NORTE"),
    ("CRED421", "POLO REDE BH"),
    ("CRED412", "POLO REDE JUIZ DE FORA"),
    ("CRED411", "POLO REDE TEOFILO OTONI"),
    ("CRED419", "POLO REDE GOV VALADARES"),
    ("CRED416", "POLO REDE MONTES CLAROS"),
    ("CRED422", "POLO REDE IPATINGA"),
    ("CRED425", "POLO REDE GUARULHOS"),
]
CRED_CODES = [code for code, _ in CRED_OPTIONS]

MM_TO_POINTS = 72 / 25.4
DATA_FMT_BR = "%d/%m/%Y"
DEFAULT_CONFIG_OUTROS = {
    "largura_mm": 90.0,
    "altura_mm": 100.0,
    "espacamento_pt": 5.0,
    "escala_fonte": 2.5,
}
DEFAULT_CONFIG_REDE = {
    "largura_mm": 150.0,
    "altura_mm": 100.0,
    "espacamento_pt": 5.0,
    "escala_fonte": 1.8,
}
DEFAULT_CONFIG_TERMICA = {
    "largura_mm": 100.0,
    "altura_mm": 80.0,
    "espacamento_pt": 4.0,
    "escala_fonte": 1.6,
}


def _apenas_numeros(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def _data_hoje_br() -> str:
    return datetime.now().strftime(DATA_FMT_BR)


@st.cache_data(show_spinner=False)
def _carregar_origens_e_cred() -> tuple[list[str], dict[str, str], str | None]:
    path = Path(__file__).with_name(PLANILHA_BASE_CRED)
    if not path.exists():
        return [], {}, f"Planilha nao encontrada: {PLANILHA_BASE_CRED}"
    if not OPENPYXL_AVAILABLE:
        return [], {}, "Biblioteca openpyxl nao encontrada para leitura da planilha."

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    origens: list[str] = []
    origem_para_cred: dict[str, str] = {}

    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        origem_raw = row[0]
        cred_raw = row[2]
        if origem_raw is None:
            continue
        origem = str(origem_raw).strip()
        if not origem:
            continue
        if origem not in origens:
            origens.append(origem)

        if cred_raw is not None and str(cred_raw).strip():
            origem_para_cred[origem] = str(cred_raw).strip().upper()

    return origens, origem_para_cred, None


def _erro_numero_obrigatorio(campo: str, valor: str, max_digitos: int | None = None) -> str | None:
    if not valor:
        return f"O campo '{campo}' e obrigatorio."
    if not valor.isdigit():
        return f"O campo '{campo}' aceita apenas numeros."
    if max_digitos and len(valor) > max_digitos:
        return f"O campo '{campo}' permite no maximo {max_digitos} digitos."
    return None


def _validar_obrigatorios_base(entradas: dict, erros: list[str]) -> None:
    if not entradas["origem"]:
        erros.append("O campo 'Origem' e obrigatorio.")
    if not entradas["destino"]:
        erros.append("O campo 'Destino' e obrigatorio.")
    if not entradas["projeto"]:
        erros.append("O campo 'Projeto' e obrigatorio.")


def _validar_campos_rede(entradas: dict, erros: list[str]) -> None:
    tecnologia = (entradas.get("tecnologia") or "").strip().upper()
    if not tecnologia:
        erros.append("O campo 'Tecnologia' e obrigatorio.")
    elif len(tecnologia) > 3:
        erros.append("O campo 'Tecnologia' permite no maximo 3 caracteres.")
    elif not re.fullmatch(r"[A-Za-z]{1,3}", tecnologia):
        erros.append("O campo 'Tecnologia' aceita apenas letras.")

    for campo, chave, limite in [
        ("Nota Fiscal", "nota_fiscal", 8),
        ("OS", "os", 10),
        ("ID FEDEX", "id_fedex", 10),
    ]:
        erro = _erro_numero_obrigatorio(campo, entradas.get(chave, ""), limite)
        if erro:
            erros.append(erro)

    if not entradas.get("numero_cred"):
        erros.append("O campo 'Numero CRED' e obrigatorio.")


def _validar_campos_padrao(entradas: dict, erros: list[str]) -> None:
    for campo, chave, limite in [
        ("Romaneio", "romaneio_sufixo", None),
        ("NR NF", "nr_nf", None),
        ("ID FEDEX", "id_fedex", 10),
    ]:
        erro = _erro_numero_obrigatorio(campo, entradas.get(chave, ""), limite)
        if erro:
            erros.append(erro)


def _validar_volume_e_layout(entradas: dict, erros: list[str]) -> None:
    err_vol = _erro_numero_obrigatorio("Volume", entradas.get("volume_total", ""), 3)
    if err_vol:
        erros.append(err_vol)
    elif int(entradas["volume_total"]) <= 0:
        erros.append("O campo 'Volume' deve ser maior que zero.")

    if entradas["largura_mm"] <= 0 or entradas["altura_mm"] <= 0:
        erros.append("Largura e altura da etiqueta devem ser maiores que zero.")
    if entradas["espacamento_linhas"] < 0:
        erros.append("Espacamento de linhas deve ser maior ou igual a zero.")
    if entradas["escala_fonte"] <= 0:
        erros.append("Escala de fonte deve ser maior que zero.")
    if entradas["ajuste_cabecalho"] < 0:
        erros.append("Ajuste cabecalho deve ser maior ou igual a zero.")
    if entradas["ajuste_rodape"] < 0:
        erros.append("Ajuste rodape deve ser maior ou igual a zero.")


def _validar_entradas(entradas: dict) -> list[str]:
    erros: list[str] = []
    _validar_obrigatorios_base(entradas, erros)
    if entradas["projeto"] == PROJETO_REDE:
        _validar_campos_rede(entradas, erros)
    else:
        _validar_campos_padrao(entradas, erros)
    _validar_volume_e_layout(entradas, erros)
    return erros


def _montar_dados_padrao(entradas: dict) -> dict:
    prefixo = PREFIXOS_ROMANEIO[entradas["projeto"]]
    romaneio = f"{prefixo}{entradas['romaneio_sufixo']}"
    total = int(entradas["volume_total"])
    total_fmt = str(total).zfill(3)
    codigo_base = _apenas_numeros(romaneio)
    data_emissao = _data_hoje_br()
    id_fedex_data = f"{entradas['id_fedex']} - {data_emissao}"

    etiquetas = []
    for i in range(1, total + 1):
        atual_fmt = str(i).zfill(3)
        volume = f"{atual_fmt}/{total_fmt}"
        etiquetas.append(
            {
                "mode": "PADRAO",
                "origem": entradas["origem"],
                "destino": entradas["destino"],
                "projeto": entradas["projeto"],
                "romaneio": romaneio,
                "nr_nf": entradas["nr_nf"],
                "id_fedex_data": id_fedex_data,
                "volume": volume,
                "codigo_barras": f"{codigo_base}{atual_fmt}{total_fmt}",
            }
        )

    return {
        "mode": "PADRAO",
        "origem": entradas["origem"],
        "destino": entradas["destino"],
        "projeto": entradas["projeto"],
        "romaneio": romaneio,
        "nr_nf": entradas["nr_nf"],
        "id_fedex_data": id_fedex_data,
        "etiquetas": etiquetas,
    }


def _montar_dados_rede(entradas: dict) -> dict:
    data_emissao = _data_hoje_br()
    total = int(entradas["volume_total"])
    total_fmt = str(total).zfill(3)
    id_fedex = entradas["id_fedex"]
    etiquetas = []
    for i in range(1, total + 1):
        atual_fmt = str(i).zfill(3)
        volume_num = f"{atual_fmt}{total_fmt}"
        etiquetas.append(
            {
                "mode": "REDE",
                "titulo": "OPERACAO REVERSA",
                "tecnologia": entradas["tecnologia"].strip().upper(),
                "origem": entradas["origem"],
                "destino": entradas["destino"],
                "numero_cred": entradas["numero_cred"],
                "nota_fiscal": entradas["nota_fiscal"],
                "data_emissao": data_emissao,
                "os": entradas["os"],
                "id_fedex": id_fedex,
                "volume": f"{atual_fmt}/{total_fmt}",
                "codigo_barras": f"{entradas['nota_fiscal']}{entradas['os']}{volume_num}",
            }
        )
    return {
        "mode": "REDE",
        "origem": entradas["origem"],
        "destino": entradas["destino"],
        "projeto": entradas["projeto"],
        "volume_total": total,
        "etiquetas": etiquetas,
    }


def _montar_dados(entradas: dict) -> dict:
    if entradas["projeto"] == PROJETO_REDE:
        return _montar_dados_rede(entradas)
    return _montar_dados_padrao(entradas)


def _layout_paginas_a4(largura_pt: float, altura_pt: float) -> dict | None:
    if not REPORTLAB_AVAILABLE or A4 is None:
        return None

    pagina_w, pagina_h = A4
    margem = 12 * MM_TO_POINTS
    gap = 4 * MM_TO_POINTS
    area_w = pagina_w - (2 * margem)
    area_h = pagina_h - (2 * margem)

    cols = int((area_w + gap) // (largura_pt + gap))
    rows = int((area_h + gap) // (altura_pt + gap))
    if cols < 1 or rows < 1:
        return None

    x0 = margem
    y0 = pagina_h - margem - altura_pt
    step_x = largura_pt + gap
    step_y = altura_pt + gap

    positions = []
    for row in range(rows):
        for col in range(cols):
            positions.append((x0 + col * step_x, y0 - row * step_y))

    return {
        "por_pagina": cols * rows,
        "positions": positions,
        "page_size": A4,
        "midia_saida": MIDIA_A4,
    }


def _layout_pagina_termica(largura_pt: float, altura_pt: float) -> dict | None:
    if not REPORTLAB_AVAILABLE:
        return None
    return {
        "por_pagina": 1,
        "positions": [(0.0, 0.0)],
        "page_size": (largura_pt, altura_pt),
        "midia_saida": MIDIA_TERMICA,
    }


def _resolver_layout_paginas(
    largura_pt: float,
    altura_pt: float,
    midia_saida: str,
) -> dict | None:
    if midia_saida == MIDIA_TERMICA:
        return _layout_pagina_termica(largura_pt, altura_pt)
    return _layout_paginas_a4(largura_pt, altura_pt)


def _desenhar_etiqueta_padrao_pdf(
    c,
    x,
    y,
    largura_pt,
    altura_pt,
    dados,
    espacamento_extra,
    escala_fonte_usuario,
    ajuste_cabecalho=0.0,
):
    draw_template_padrao(
        c=c,
        barcode_module=code128,
        x=x,
        y=y,
        largura_pt=largura_pt,
        altura_pt=altura_pt,
        dados=dados,
        mm_to_points=MM_TO_POINTS,
        app_name=APP_NAME,
        espacamento_extra=espacamento_extra,
        escala_fonte_usuario=escala_fonte_usuario,
        ajuste_cabecalho=ajuste_cabecalho,
    )


def _desenhar_etiqueta_rede_pdf(
    c,
    x,
    y,
    largura_pt,
    altura_pt,
    dados,
    espacamento_extra,
    escala_fonte_usuario,
    ajuste_cabecalho=0.0,
    ajuste_rodape=0.0,
):
    draw_template_rede(
        c=c,
        barcode_module=code128,
        x=x,
        y=y,
        largura_pt=largura_pt,
        altura_pt=altura_pt,
        dados=dados,
        mm_to_points=MM_TO_POINTS,
        espacamento_extra=espacamento_extra,
        escala_fonte_usuario=escala_fonte_usuario,
        ajuste_cabecalho=ajuste_cabecalho,
        ajuste_rodape=ajuste_rodape,
    )


def _gerar_pdf_bytes(
    dados_lote: dict,
    largura_mm: float,
    altura_mm: float,
    espacamento_extra: float,
    escala_fonte_usuario: float,
    ajuste_cabecalho: float = 0.0,
    ajuste_rodape: float = 0.0,
    midia_saida: str = MIDIA_A4,
) -> bytes:
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("Biblioteca reportlab nao encontrada.")

    largura_pt = largura_mm * MM_TO_POINTS
    altura_pt = altura_mm * MM_TO_POINTS
    layout = _resolver_layout_paginas(largura_pt, altura_pt, midia_saida)
    if not layout:
        raise ValueError("Com esse tamanho de etiqueta nao cabe nenhuma unidade na folha A4.")

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=layout["page_size"])
    c.setTitle(APP_NAME)

    for i, etiqueta in enumerate(dados_lote["etiquetas"]):
        slot = i % layout["por_pagina"]
        if i > 0 and slot == 0:
            c.showPage()
        x, y = layout["positions"][slot]

        if etiqueta["mode"] == "REDE":
            _desenhar_etiqueta_rede_pdf(
                c,
                x,
                y,
                largura_pt,
                altura_pt,
                etiqueta,
                espacamento_extra,
                escala_fonte_usuario,
                ajuste_cabecalho,
                ajuste_rodape,
            )
        else:
            _desenhar_etiqueta_padrao_pdf(
                c,
                x,
                y,
                largura_pt,
                altura_pt,
                etiqueta,
                espacamento_extra,
                escala_fonte_usuario,
                ajuste_cabecalho,
            )

    c.save()
    return buf.getvalue()


def _init_state() -> None:
    if "resultado" not in st.session_state:
        st.session_state["resultado"] = None
    if "pdf_bytes" not in st.session_state:
        st.session_state["pdf_bytes"] = None
    if "erros" not in st.session_state:
        st.session_state["erros"] = []
    if "origem_rede_prev" not in st.session_state:
        st.session_state["origem_rede_prev"] = ""
    if "cred_code" not in st.session_state:
        st.session_state["cred_code"] = ""
    if "cfg_contexto_prev" not in st.session_state:
        st.session_state["cfg_contexto_prev"] = ""
    if "cfg_midia_saida" not in st.session_state:
        st.session_state["cfg_midia_saida"] = MIDIA_A4


def _config_padrao_contexto(is_rede: bool, midia_saida: str) -> dict[str, float]:
    if midia_saida == MIDIA_TERMICA:
        return DEFAULT_CONFIG_TERMICA
    return DEFAULT_CONFIG_REDE if is_rede else DEFAULT_CONFIG_OUTROS


def _aplicar_config_padrao_por_contexto(is_rede: bool, midia_saida: str) -> None:
    modo_cfg = PROJETO_REDE if is_rede else "OUTROS"
    contexto_cfg = f"{modo_cfg}|{midia_saida}"
    cfg_padrao = _config_padrao_contexto(is_rede, midia_saida)
    campos_cfg = [
        ("cfg_largura_mm", "largura_mm"),
        ("cfg_altura_mm", "altura_mm"),
        ("cfg_espacamento_pt", "espacamento_pt"),
        ("cfg_escala_fonte", "escala_fonte"),
    ]

    if st.session_state.get("cfg_contexto_prev") != contexto_cfg:
        for chave_state, chave_cfg in campos_cfg:
            st.session_state[chave_state] = cfg_padrao[chave_cfg]
        st.session_state["cfg_contexto_prev"] = contexto_cfg
        return

    for chave_state, chave_cfg in campos_cfg:
        if chave_state not in st.session_state:
            st.session_state[chave_state] = cfg_padrao[chave_cfg]


def _render_secao_base(origens: list[str]) -> tuple[str, str, str, bool]:
    st.subheader("Origem / Destino / Projeto")
    c1, c2, c3 = st.columns(3)
    origem = c1.selectbox("Origem *", [""] + origens, index=0)
    destino = c2.selectbox("Destino *", [""] + DESTINOS, index=0)
    projeto = c3.selectbox("Projeto *", [""] + PROJETOS, index=0)
    return origem, destino, projeto, projeto == PROJETO_REDE


def _render_secao_rede(origem: str, origem_para_cred: dict[str, str]) -> dict[str, str]:
    st.subheader("Projeto REDE")
    cred_sugerido = origem_para_cred.get(origem, "") if origem else ""

    if st.session_state["origem_rede_prev"] != origem:
        st.session_state["origem_rede_prev"] = origem
        st.session_state["cred_code"] = cred_sugerido if cred_sugerido in CRED_CODES else ""

    r1, r2, r3, r4 = st.columns(4)
    tecnologia = r1.text_input("Tecnologia * (texto, max 3)", max_chars=3).strip()
    nota_fiscal = r2.text_input("Nota Fiscal * (max 8)", max_chars=8).strip()
    os_num = r3.text_input("OS * (max 10)", max_chars=10).strip()
    id_fedex = r4.text_input("ID FEDEX * (max 10)", max_chars=10).strip()

    r5, r6, r7 = st.columns([1, 2, 1])
    r5.text_input("Data Emissao", value=_data_hoje_br(), disabled=True)
    numero_cred = r6.selectbox("Numero CRED *", [""] + CRED_CODES, key="cred_code")
    volume_total = r7.text_input("Volume *", max_chars=3).strip()

    if cred_sugerido:
        st.caption(f"CRED sugerido pela Origem: {cred_sugerido}")

    return {
        "tecnologia": tecnologia,
        "nota_fiscal": nota_fiscal,
        "os": os_num,
        "numero_cred": numero_cred,
        "romaneio_sufixo": "",
        "nr_nf": "",
        "id_fedex": id_fedex,
        "volume_total": volume_total,
    }


def _render_secao_outros(projeto: str) -> dict[str, str]:
    st.subheader("Outros Projetos")
    prefixo = PREFIXOS_ROMANEIO.get(projeto, "")
    o1, o2 = st.columns([1, 2])
    o1.text_input("Prefixo Romaneio", value=prefixo, disabled=True)
    romaneio_sufixo = o2.text_input("Romaneio (numeros apos /) *", max_chars=20).strip()
    o3, o4 = st.columns(2)
    nr_nf = o3.text_input("NR NF *", max_chars=20).strip()
    id_fedex = o4.text_input("ID FEDEX *", max_chars=10).strip()
    volume_total = st.text_input("Volume (qtd total de etiquetas) *", max_chars=3).strip()

    return {
        "tecnologia": "",
        "nota_fiscal": "",
        "os": "",
        "numero_cred": "",
        "romaneio_sufixo": romaneio_sufixo,
        "nr_nf": nr_nf,
        "id_fedex": id_fedex,
        "volume_total": volume_total,
    }


def _render_secao_configuracao(is_rede: bool) -> dict[str, float | str]:
    st.subheader("Configuracao da Etiqueta")
    midia_saida = st.selectbox(
        "Midia de impressao",
        [MIDIA_A4, MIDIA_TERMICA],
        key="cfg_midia_saida",
    )
    _aplicar_config_padrao_por_contexto(is_rede, midia_saida)
    if midia_saida == MIDIA_TERMICA:
        st.caption(
            "Para Zebra/Argox use este modo. O PDF sai com 1 etiqueta por pagina, "
            "no tamanho exato configurado."
        )
    s1, s2, s3, s4 = st.columns(4)
    largura_mm = float(s1.number_input("Largura (mm)", min_value=1.0, key="cfg_largura_mm"))
    altura_mm = float(s2.number_input("Altura (mm)", min_value=1.0, key="cfg_altura_mm"))
    espacamento_linhas = float(
        s3.number_input("Espacamento linhas (pt)", min_value=0.0, key="cfg_espacamento_pt")
    )
    escala_fonte = float(
        s4.number_input("Escala de fonte", min_value=0.1, key="cfg_escala_fonte")
    )
    s5, s6 = st.columns(2)
    ajuste_cabecalho = float(
        s5.number_input("Ajuste cabecalho (pt)", min_value=0.0, value=3.0, step=0.5)
    )
    ajuste_rodape = float(
        s6.number_input("Ajuste rodape (pt)", min_value=0.0, value=3.0, step=0.5)
    )
    return {
        "midia_saida": midia_saida,
        "largura_mm": largura_mm,
        "altura_mm": altura_mm,
        "espacamento_linhas": espacamento_linhas,
        "escala_fonte": escala_fonte,
        "ajuste_cabecalho": ajuste_cabecalho,
        "ajuste_rodape": ajuste_rodape,
    }


def _montar_entradas_formulario(
    origem: str,
    destino: str,
    projeto: str,
    campos_projeto: dict[str, str],
    config_layout: dict[str, float],
) -> dict:
    return {
        "origem": origem,
        "destino": destino,
        "projeto": projeto,
        "tecnologia": campos_projeto["tecnologia"],
        "nota_fiscal": _apenas_numeros(campos_projeto["nota_fiscal"]),
        "os": _apenas_numeros(campos_projeto["os"]),
        "numero_cred": (campos_projeto["numero_cred"] or "").strip().upper(),
        "romaneio_sufixo": _apenas_numeros(campos_projeto["romaneio_sufixo"]),
        "nr_nf": _apenas_numeros(campos_projeto["nr_nf"]),
        "id_fedex": _apenas_numeros(campos_projeto["id_fedex"]),
        "volume_total": _apenas_numeros(campos_projeto["volume_total"]),
        **config_layout,
    }


def _processar_geracao(entradas: dict) -> None:
    st.session_state["erros"] = _validar_entradas(entradas)
    st.session_state["resultado"] = None
    st.session_state["pdf_bytes"] = None
    if st.session_state["erros"]:
        return

    dados = _montar_dados(entradas)
    st.session_state["resultado"] = {"dados": dados, "config": entradas}
    if not REPORTLAB_AVAILABLE:
        return

    try:
        st.session_state["pdf_bytes"] = _gerar_pdf_bytes(
            dados,
            largura_mm=entradas["largura_mm"],
            altura_mm=entradas["altura_mm"],
            espacamento_extra=entradas["espacamento_linhas"],
            escala_fonte_usuario=entradas["escala_fonte"],
            ajuste_cabecalho=entradas["ajuste_cabecalho"],
            ajuste_rodape=entradas["ajuste_rodape"],
            midia_saida=entradas["midia_saida"],
        )
    except Exception as exc:
        st.session_state["erros"].append(str(exc))
        st.session_state["resultado"] = None


def _render_erros() -> None:
    for erro in st.session_state.get("erros", []):
        st.error(erro)


def _render_preview_rede(etiquetas: list[dict]) -> None:
    e = etiquetas[0]
    volumes_preview = [item["volume"] for item in etiquetas[:40]]
    codigos_preview = [f"{item['volume']} -> {item['codigo_barras']}" for item in etiquetas[:40]]
    if len(etiquetas) > 40:
        volumes_preview.append(f"... (mostrando 40 de {len(etiquetas)} etiquetas)")
        codigos_preview.append(f"... (mostrando 40 de {len(etiquetas)} etiquetas)")
    st.code(
        (
            f"Titulo: {e['titulo']}\n"
            f"Tecnologia: {e['tecnologia']}\n"
            f"Origem: {e['origem']}\n"
            f"Destino: {e['destino']}\n"
            f"Numero CRED: {e['numero_cred']}\n"
            f"Nota Fiscal: {e['nota_fiscal']}\n"
            f"Data Emissao: {e['data_emissao']}\n"
            f"OS: {e['os']}\n"
            f"ID FEDEX: {e['id_fedex']}\n"
            f"Quantidade de etiquetas: {len(etiquetas)}\n"
            "Volumes:\n"
            + "\n".join(volumes_preview)
            + "\n\nVolumes / Codigos:\n"
            + "\n".join(codigos_preview)
        ),
        language="text",
    )


def _render_preview_padrao(dados: dict, etiquetas: list[dict]) -> None:
    preview = [f"{e['volume']} -> {e['codigo_barras']}" for e in etiquetas[:120]]
    if len(etiquetas) > 120:
        preview.append(f"... (mostrando 120 de {len(etiquetas)} etiquetas)")
    st.code(
        (
            f"Origem: {dados['origem']}\n"
            f"Destino: {dados['destino']}\n"
            f"Projeto: {dados['projeto']}\n"
            f"Romaneio: {dados['romaneio']}\n"
            f"NR NF: {dados['nr_nf']}\n"
            f"ID FEDEX: {dados['id_fedex_data']}\n"
            f"Quantidade de etiquetas: {len(etiquetas)}\n\n"
            "Volumes / Codigos de barras:\n"
            + "\n".join(preview)
        ),
        language="text",
    )


def _render_resumo_layout(resultado: dict, quantidade_etiquetas: int) -> None:
    layout = _resolver_layout_paginas(
        resultado["config"]["largura_mm"] * MM_TO_POINTS,
        resultado["config"]["altura_mm"] * MM_TO_POINTS,
        resultado["config"]["midia_saida"],
    )
    if not layout:
        return
    por_folha = layout["por_pagina"]
    folhas = (quantidade_etiquetas + por_folha - 1) // por_folha
    if resultado["config"]["midia_saida"] == MIDIA_TERMICA:
        st.info(f"Modo termico 100x80 | 1 etiqueta por pagina | Etiquetas estimadas: {folhas}")
        return
    st.info(f"Etiquetas por folha A4: {por_folha} | Folhas estimadas: {folhas}")


def _render_download_pdf() -> None:
    if not REPORTLAB_AVAILABLE:
        st.warning("Instale reportlab para gerar PDF: pip install reportlab")
        return
    if st.session_state.get("pdf_bytes"):
        st.download_button(
            "Baixar etiqueta(s) em PDF",
            data=st.session_state["pdf_bytes"],
            file_name=f"etiqueta_{datetime.now():%Y%m%d_%H%M%S}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )


def _render_resultado() -> None:
    resultado = st.session_state.get("resultado")
    if not resultado:
        return

    dados = resultado["dados"]
    etiquetas = dados["etiquetas"]
    st.success(f"Etiquetas validadas. Quantidade: {len(etiquetas)}")
    _render_resumo_layout(resultado, len(etiquetas))

    st.markdown("**Preview**")
    if dados["mode"] == "REDE":
        _render_preview_rede(etiquetas)
    else:
        _render_preview_padrao(dados, etiquetas)
    _render_download_pdf()


def main() -> None:
    st.set_page_config(page_title=APP_NAME, layout="wide")
    _init_state()

    origens, origem_para_cred, aviso_planilha = _carregar_origens_e_cred()
    st.title(APP_NAME)
    if aviso_planilha:
        st.warning(aviso_planilha)

    origem, destino, projeto, is_rede = _render_secao_base(origens)
    campos_projeto = (
        _render_secao_rede(origem, origem_para_cred)
        if is_rede
        else _render_secao_outros(projeto)
    )
    config_layout = _render_secao_configuracao(is_rede)

    if st.button("Gerar etiqueta(s)", type="primary", use_container_width=True):
        entradas = _montar_entradas_formulario(
            origem,
            destino,
            projeto,
            campos_projeto,
            config_layout,
        )
        _processar_geracao(entradas)

    _render_erros()
    _render_resultado()


if __name__ == "__main__":
    main()
