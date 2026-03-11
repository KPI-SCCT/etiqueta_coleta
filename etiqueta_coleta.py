import re
import tempfile
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from etiqueta_layout_engine import draw_template_padrao, draw_template_rede

try:
    from reportlab.graphics.barcode import code128
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    REPORTLAB_AVAILABLE = True
except Exception:
    code128 = None
    A4 = None
    canvas = None
    REPORTLAB_AVAILABLE = False

try:
    import win32api
    import win32print

    PYWIN32_AVAILABLE = True
except Exception:
    PYWIN32_AVAILABLE = False

try:
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except Exception:
    load_workbook = None
    OPENPYXL_AVAILABLE = False


APP_NAME = "COLETA"
PROJETO_REDE = "REDE"
PLANILHA_BASE_CRED = "bases padrÃ£o + cred.xlsx"
FONT_FAMILY_SEGOE_UI = "Segoe UI"
MSG_CAMPO_OBRIGATORIO = "Campo obrigatorio"
MSG_CAMPO_INVALIDO = "Campo invalido"
MSG_TAMANHO_INVALIDO = "Tamanho invalido"
MSG_AJUSTE_INVALIDO = "Ajuste invalido"
DESTINOS = ["CTDI DO BR - SP", "FLEXTRONIC", "FEDEX CAJAMAR - SP", "DHL LOUVEIRA - SP"]
PROJETOS = [
    "CIELO - POS",
    "CIELO - TEF",
    "CIELO - TRANSF",
    "FISERV",
    "MOOZ",
    "STONE",
    "PICPAY",
    "PAGBANK",
    "CTRENDS",
    "C6BANK",
    "ADYEN",
    "CLOUDWALK",
    PROJETO_REDE,
]
PREFIXOS_ROMANEIO = {
    "CIELO - POS": "1.2/",
    "CIELO - TEF": "2.2/",
    "CIELO - TRANSF": "1.3/",
    "FISERV": "34.3/",
    "MOOZ": "42.3/",
    "STONE": "41.3/",
    "PICPAY": "49.3/",
    "PAGBANK": "53.3/",
    "CTRENDS": "39.3/",
    "C6BANK": "43.3/",
    "ADYEN": "45.3/",
    "CLOUDWALK": "40.3/",
    PROJETO_REDE: "51.2/",
}
CRED_OPTIONS = [
    ("CRED369", "POLO REDE PONTA GROSSA"),
    ("CRED385", "POLO REDE SJ DOS CAMPOS"),
    ("CRED368", "POLO REDE SANTOS"),
    ("CRED372", "POLO REDE CUR PINHAIS"),
    ("CRED382", "POLO REDE LONDRINA"),
    ("CRED371", "POLO REDE CURITIBA"),
    ("CRED370", "POLO REDE MARINGA"),
    ("CRED384", "POLO REDE PINDA"),
    ("CRED383", "POLO REDE CASCAVEL"),
    ("CRED408", "POLO REDE FORTALEZA"),
    ("CRED409", "POLO REDE JUAZ DO NORTE"),
    ("CRED421", "POLO REDE BH"),
    ("CRED412", "POLO REDE JUIZ DE FORA"),
    ("CRED411", "POLO REDE TEOFILO OTONI"),
    ("CRED419", "POLO REDE GOV VALADARES"),
    ("CRED416", "POLO REDE MONTES CLAROS"),
    ("CRED422", "POLO REDE IPATINGA"),
    ("CRED425", "POLO REDE GUARULHOS"),
]
CRED_CODES = [codigo for codigo, _ in CRED_OPTIONS]
PRINTER_DEFAULT_LABEL = "Padrao do sistema"
MM_TO_POINTS = 72 / 25.4
DEFAULT_CONFIG_OUTROS = {
    "largura_mm": "90",
    "altura_mm": "100",
    "espacamento_pt": "5",
    "escala_fonte": "2.5",
}
DEFAULT_CONFIG_REDE = {
    "largura_mm": "150",
    "altura_mm": "100",
    "espacamento_pt": "5",
    "escala_fonte": "1.8",
}


def _carregar_origens_e_cred() -> tuple[list[str], dict[str, str], str | None]:
    path = Path(__file__).with_name(PLANILHA_BASE_CRED)
    if not path.exists():
        return [], {}, f"Planilha nao encontrada: {PLANILHA_BASE_CRED}"
    if not OPENPYXL_AVAILABLE:
        return [], {}, "Biblioteca openpyxl nao encontrada para ler a planilha."

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    origens: list[str] = []
    origem_para_cred: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        origem_raw = row[0]
        cred_raw = row[2]
        if origem_raw is None:
            continue
        origem = str(origem_raw).strip()
        if not origem:
            continue
        if origem not in origens:
            origens.append(origem)

        if cred_raw is not None and str(cred_raw).strip():
            origem_para_cred[origem] = str(cred_raw).strip().upper()

    return origens, origem_para_cred, None


def _label_por_codigo_cred(codigo: str) -> str:
    return (codigo or "").strip().upper()


class EtiquetaColetaApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title(APP_NAME)
        self.root.minsize(860, 680)

        self.romaneio_prefixo_var = tk.StringVar()
        self.romaneio_sufixo_var = tk.StringVar()
        self.nr_nf_var = tk.StringVar()
        self.id_fedex_var = tk.StringVar()
        self.volume_qtd_var = tk.StringVar()
        self.tecnologia_var = tk.StringVar()
        self.nota_fiscal_var = tk.StringVar()
        self.os_var = tk.StringVar()
        self.codigo_barras_var = tk.StringVar()
        self.etiqueta_largura_var = tk.StringVar(value=DEFAULT_CONFIG_OUTROS["largura_mm"])
        self.etiqueta_altura_var = tk.StringVar(value=DEFAULT_CONFIG_OUTROS["altura_mm"])
        self.espacamento_linhas_var = tk.StringVar(value=DEFAULT_CONFIG_OUTROS["espacamento_pt"])
        self.escala_fonte_var = tk.StringVar(value=DEFAULT_CONFIG_OUTROS["escala_fonte"])
        self.ajuste_cabecalho_var = tk.StringVar(value="3.0")
        self.ajuste_rodape_var = tk.StringVar(value="3.0")
        self.impressora_var = tk.StringVar(value=PRINTER_DEFAULT_LABEL)
        self.origens, self.origem_para_cred, self.aviso_planilha = _carregar_origens_e_cred()

        self.vcmd_digitos = (self.root.register(self._validar_digitos), "%P")
        self.vcmd_volume = (self.root.register(self._validar_volume), "%P")
        self.vcmd_texto_3 = (self.root.register(self._validar_texto_3), "%P")
        self.vcmd_nf_8 = (self.root.register(self._validar_nf_8), "%P")
        self.vcmd_os_10 = (self.root.register(self._validar_os_10), "%P")
        self.vcmd_decimal = (self.root.register(self._validar_decimal), "%P")

        self._montar_layout()
        self._carregar_impressoras()
        self._atualizar_prefixo_romaneio()
        self._atualizar_modo_projeto()
        if self.aviso_planilha:
            messagebox.showwarning("Aviso", self.aviso_planilha)

    def _montar_layout(self) -> None:
        frame = ttk.Frame(self.root, padding=14)
        frame.grid(sticky="nsew")
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        for col in range(4):
            frame.columnconfigure(col, weight=1)

        ttk.Label(
            frame, text=APP_NAME, font=(FONT_FAMILY_SEGOE_UI, 13, "bold")
        ).grid(row=0, column=0, columnspan=4, sticky="ew", pady=(0, 10))

        ttk.Label(
            frame,
            text="Sessao 1 - Origem / Destino / Projeto",
            font=(FONT_FAMILY_SEGOE_UI, 10, "bold"),
        ).grid(
            row=1, column=0, columnspan=4, sticky="w", pady=(6, 4)
        )

        ttk.Label(frame, text="Origem:").grid(row=2, column=0, sticky="w")
        origem_frame = ttk.Frame(frame)
        origem_frame.grid(row=3, column=0, sticky="nsew", padx=(0, 8))
        origem_frame.columnconfigure(0, weight=1)
        self.lb_origem = tk.Listbox(origem_frame, height=7, exportselection=False)
        self.lb_origem.grid(row=0, column=0, sticky="nsew")
        sb_origem = ttk.Scrollbar(origem_frame, orient="vertical", command=self.lb_origem.yview)
        sb_origem.grid(row=0, column=1, sticky="ns")
        self.lb_origem.config(yscrollcommand=sb_origem.set)
        for origem in (self.origens or ["Sem origem carregada"]):
            self.lb_origem.insert(tk.END, origem)
        self.lb_origem.selection_set(0)
        self.lb_origem.bind("<<ListboxSelect>>", self._on_origem_change)

        ttk.Label(frame, text="Destino:").grid(row=2, column=1, sticky="w")
        self.lb_destino = tk.Listbox(frame, height=7, exportselection=False)
        self.lb_destino.grid(row=3, column=1, sticky="nsew", padx=(0, 8))
        for destino in DESTINOS:
            self.lb_destino.insert(tk.END, destino)
        self.lb_destino.selection_set(0)

        ttk.Label(frame, text="Projeto:").grid(row=2, column=2, sticky="w")
        self.lb_projeto = tk.Listbox(frame, height=7, exportselection=False)
        self.lb_projeto.grid(row=3, column=2, columnspan=2, sticky="nsew")
        for projeto in PROJETOS:
            self.lb_projeto.insert(tk.END, projeto)
        self.lb_projeto.selection_set(0)
        self.lb_projeto.bind("<<ListboxSelect>>", self._on_projeto_change)

        ttk.Label(
            frame,
            text="Sessao 2 - Campos do Projeto",
            font=(FONT_FAMILY_SEGOE_UI, 10, "bold"),
        ).grid(
            row=4, column=0, columnspan=4, sticky="w", pady=(10, 4)
        )

        self.frame_nao_rede = ttk.Frame(frame)
        self.frame_nao_rede.grid(row=5, column=0, columnspan=4, sticky="ew")
        for col in range(4):
            self.frame_nao_rede.columnconfigure(col, weight=1)

        ttk.Label(self.frame_nao_rede, text="Romaneio:").grid(row=0, column=0, sticky="w", pady=(2, 0))
        romaneio_frame = ttk.Frame(self.frame_nao_rede)
        romaneio_frame.grid(row=1, column=0, columnspan=2, sticky="w", padx=(0, 8))
        ttk.Label(
            romaneio_frame,
            textvariable=self.romaneio_prefixo_var,
            width=7,
            anchor="w",
            font=(FONT_FAMILY_SEGOE_UI, 10, "bold"),
        ).pack(side=tk.LEFT)
        ttk.Entry(
            romaneio_frame,
            textvariable=self.romaneio_sufixo_var,
            validate="key",
            validatecommand=self.vcmd_digitos,
            width=24,
        ).pack(side=tk.LEFT, padx=(6, 0))

        ttk.Label(self.frame_nao_rede, text="NR NF:").grid(row=0, column=2, sticky="w", pady=(2, 0))
        ttk.Entry(
            self.frame_nao_rede,
            textvariable=self.nr_nf_var,
            validate="key",
            validatecommand=self.vcmd_digitos,
        ).grid(row=1, column=2, columnspan=2, sticky="ew")

        ttk.Label(self.frame_nao_rede, text="ID FEDEX:").grid(row=2, column=0, sticky="w", pady=(10, 0))
        ttk.Entry(
            self.frame_nao_rede,
            textvariable=self.id_fedex_var,
            validate="key",
            validatecommand=self.vcmd_os_10,
        ).grid(row=3, column=0, columnspan=2, sticky="ew", padx=(0, 8))

        ttk.Label(self.frame_nao_rede, text="Volume (qtd total):").grid(
            row=2, column=2, sticky="w", pady=(10, 0)
        )
        volume_frame = ttk.Frame(self.frame_nao_rede)
        volume_frame.grid(row=3, column=2, columnspan=2, sticky="w")
        ttk.Entry(
            volume_frame,
            textvariable=self.volume_qtd_var,
            width=6,
            validate="key",
            validatecommand=self.vcmd_volume,
        ).pack(side=tk.LEFT)
        ttk.Label(volume_frame, text="(max 3 digitos)", font=(FONT_FAMILY_SEGOE_UI, 9)).pack(
            side=tk.LEFT, padx=(8, 0)
        )

        self.frame_rede = ttk.Frame(frame)
        self.frame_rede.grid(row=5, column=0, columnspan=4, sticky="ew")
        for col in range(4):
            self.frame_rede.columnconfigure(col, weight=1)

        ttk.Label(self.frame_rede, text="Tecnologia:").grid(row=0, column=0, sticky="w", pady=(2, 0))
        ttk.Entry(
            self.frame_rede,
            textvariable=self.tecnologia_var,
            validate="key",
            validatecommand=self.vcmd_texto_3,
            width=8,
        ).grid(row=1, column=0, sticky="w", padx=(0, 8))

        ttk.Label(self.frame_rede, text="Nota Fiscal:").grid(row=0, column=1, sticky="w", pady=(2, 0))
        ttk.Entry(
            self.frame_rede,
            textvariable=self.nota_fiscal_var,
            validate="key",
            validatecommand=self.vcmd_nf_8,
            width=14,
        ).grid(row=1, column=1, sticky="w", padx=(0, 8))

        ttk.Label(self.frame_rede, text="OS:").grid(row=0, column=2, sticky="w", pady=(2, 0))
        ttk.Entry(
            self.frame_rede,
            textvariable=self.os_var,
            validate="key",
            validatecommand=self.vcmd_os_10,
            width=14,
        ).grid(row=1, column=2, sticky="w", padx=(0, 8))

        ttk.Label(self.frame_rede, text="ID FEDEX:").grid(row=0, column=3, sticky="w", pady=(2, 0))
        ttk.Entry(
            self.frame_rede,
            textvariable=self.id_fedex_var,
            validate="key",
            validatecommand=self.vcmd_os_10,
            width=14,
        ).grid(row=1, column=3, sticky="w")

        ttk.Label(self.frame_rede, text="Numero CRED:").grid(row=2, column=0, sticky="w", pady=(10, 0))
        cred_frame = ttk.Frame(self.frame_rede)
        cred_frame.grid(row=3, column=0, columnspan=2, sticky="nsew", padx=(0, 8))
        cred_frame.columnconfigure(0, weight=1)
        self.lb_cred = tk.Listbox(cred_frame, height=6, exportselection=False)
        self.lb_cred.grid(row=0, column=0, sticky="ew")
        sb_cred = ttk.Scrollbar(cred_frame, orient="vertical", command=self.lb_cred.yview)
        sb_cred.grid(row=0, column=1, sticky="ns")
        self.lb_cred.config(yscrollcommand=sb_cred.set)
        for cred_code in CRED_CODES:
            self.lb_cred.insert(tk.END, cred_code)
        if CRED_CODES:
            self.lb_cred.selection_set(0)

        ttk.Label(self.frame_rede, text="Volume (qtd total):").grid(row=2, column=2, sticky="w", pady=(10, 0))
        ttk.Entry(
            self.frame_rede,
            textvariable=self.volume_qtd_var,
            validate="key",
            validatecommand=self.vcmd_volume,
            width=14,
        ).grid(row=3, column=2, sticky="w", padx=(0, 8))

        ttk.Label(self.frame_rede, text="Data Emissao:").grid(row=2, column=3, sticky="w", pady=(10, 0))
        self.rede_data_emissao_entry = ttk.Entry(self.frame_rede, state="readonly", width=14)
        self.rede_data_emissao_entry.grid(row=3, column=3, sticky="w")

        ttk.Label(
            frame,
            text="Sessao 3 - Configuracao da Etiqueta",
            font=(FONT_FAMILY_SEGOE_UI, 10, "bold"),
        ).grid(
            row=6, column=0, columnspan=4, sticky="w", pady=(12, 4)
        )
        tamanho_frame = ttk.Frame(frame)
        tamanho_frame.grid(row=7, column=0, columnspan=4, sticky="w")
        ttk.Label(tamanho_frame, text="Largura").pack(side=tk.LEFT)
        ttk.Entry(
            tamanho_frame,
            textvariable=self.etiqueta_largura_var,
            width=8,
            validate="key",
            validatecommand=self.vcmd_decimal,
        ).pack(side=tk.LEFT, padx=(6, 12))
        ttk.Label(tamanho_frame, text="Altura").pack(side=tk.LEFT)
        ttk.Entry(
            tamanho_frame,
            textvariable=self.etiqueta_altura_var,
            width=8,
            validate="key",
            validatecommand=self.vcmd_decimal,
        ).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Label(tamanho_frame, text="Espacamento (pt)").pack(side=tk.LEFT, padx=(12, 0))
        ttk.Entry(
            tamanho_frame,
            textvariable=self.espacamento_linhas_var,
            width=6,
            validate="key",
            validatecommand=self.vcmd_decimal,
        ).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Label(tamanho_frame, text="Escala fonte").pack(side=tk.LEFT, padx=(12, 0))
        ttk.Entry(
            tamanho_frame,
            textvariable=self.escala_fonte_var,
            width=6,
            validate="key",
            validatecommand=self.vcmd_decimal,
        ).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Label(tamanho_frame, text="Ajuste cabecalho (pt)").pack(side=tk.LEFT, padx=(12, 0))
        ttk.Entry(
            tamanho_frame,
            textvariable=self.ajuste_cabecalho_var,
            width=6,
            validate="key",
            validatecommand=self.vcmd_decimal,
        ).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Label(tamanho_frame, text="Ajuste rodape (pt)").pack(side=tk.LEFT, padx=(12, 0))
        ttk.Entry(
            tamanho_frame,
            textvariable=self.ajuste_rodape_var,
            width=6,
            validate="key",
            validatecommand=self.vcmd_decimal,
        ).pack(side=tk.LEFT, padx=(6, 0))

        ttk.Label(frame, text="Impressora:").grid(row=8, column=0, sticky="w", pady=(12, 0))
        impressora_frame = ttk.Frame(frame)
        impressora_frame.grid(row=9, column=0, columnspan=4, sticky="ew")
        impressora_frame.columnconfigure(0, weight=1)
        self.cmb_impressora = ttk.Combobox(
            impressora_frame,
            textvariable=self.impressora_var,
            state="readonly",
        )
        self.cmb_impressora.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        ttk.Button(
            impressora_frame, text="Atualizar impressoras", command=self._carregar_impressoras
        ).grid(row=0, column=1, sticky="e")

        ttk.Label(frame, text="Codigo de barras:").grid(
            row=10, column=0, sticky="w", pady=(12, 0)
        )
        ttk.Entry(
            frame,
            textvariable=self.codigo_barras_var,
            state="readonly",
            font=("Consolas", 10),
        ).grid(row=11, column=0, columnspan=4, sticky="ew")

        ttk.Label(frame, text="Preview da etiqueta:").grid(
            row=12, column=0, sticky="w", pady=(12, 0)
        )
        self.txt_preview = tk.Text(frame, height=10, state="disabled", wrap="word")
        self.txt_preview.grid(row=13, column=0, columnspan=4, sticky="nsew")
        frame.rowconfigure(13, weight=1)

        botoes_frame = ttk.Frame(frame)
        botoes_frame.grid(row=14, column=0, columnspan=4, sticky="e", pady=(12, 0))
        ttk.Button(botoes_frame, text="Gerar codigo", command=self.gerar_codigo).pack(
            side=tk.LEFT, padx=6
        )
        ttk.Button(botoes_frame, text="Salvar PDF", command=self.salvar_pdf).pack(
            side=tk.LEFT, padx=6
        )
        ttk.Button(botoes_frame, text="Imprimir etiqueta", command=self.imprimir).pack(
            side=tk.LEFT, padx=6
        )

    @staticmethod
    def _validar_digitos(valor: str) -> bool:
        return valor == "" or valor.isdigit()

    @staticmethod
    def _validar_volume(valor: str) -> bool:
        return valor == "" or (valor.isdigit() and len(valor) <= 3)

    @staticmethod
    def _validar_texto_3(valor: str) -> bool:
        if valor == "":
            return True
        return bool(re.fullmatch(r"[A-Za-z]{1,3}", valor))

    @staticmethod
    def _validar_nf_8(valor: str) -> bool:
        return valor == "" or (valor.isdigit() and len(valor) <= 8)

    @staticmethod
    def _validar_os_10(valor: str) -> bool:
        return valor == "" or (valor.isdigit() and len(valor) <= 10)

    @staticmethod
    def _validar_decimal(valor: str) -> bool:
        if valor == "":
            return True
        valor = valor.replace(",", ".")
        try:
            float(valor)
            return True
        except ValueError:
            return False

    def _on_origem_change(self, _: tk.Event) -> None:
        self._preencher_cred_por_origem()

    def _on_projeto_change(self, _: tk.Event) -> None:
        self._atualizar_prefixo_romaneio()
        self._atualizar_modo_projeto()
        self._preencher_cred_por_origem()

    def _atualizar_prefixo_romaneio(self) -> None:
        projeto = self._valor_listbox(self.lb_projeto, "Projeto", exibir_erro=False)
        if projeto:
            self.romaneio_prefixo_var.set(PREFIXOS_ROMANEIO[projeto])
        else:
            self.romaneio_prefixo_var.set("")

    def _modo_rede_ativo(self) -> bool:
        projeto = self._valor_listbox(self.lb_projeto, "Projeto", exibir_erro=False)
        return projeto == PROJETO_REDE

    def _atualizar_modo_projeto(self) -> None:
        data_hoje = datetime.now().strftime("%d/%m/%Y")
        self.rede_data_emissao_entry.config(state="normal")
        self.rede_data_emissao_entry.delete(0, tk.END)
        self.rede_data_emissao_entry.insert(0, data_hoje)
        self.rede_data_emissao_entry.config(state="readonly")

        modo_rede = self._modo_rede_ativo()
        self._aplicar_config_padrao_por_projeto(modo_rede)

        if modo_rede:
            self.frame_nao_rede.grid_remove()
            self.frame_rede.grid()
            self.codigo_barras_var.set("-")
        else:
            self.frame_rede.grid_remove()
            self.frame_nao_rede.grid()

    def _aplicar_config_padrao_por_projeto(self, modo_rede: bool) -> None:
        cfg = DEFAULT_CONFIG_REDE if modo_rede else DEFAULT_CONFIG_OUTROS
        self.etiqueta_largura_var.set(cfg["largura_mm"])
        self.etiqueta_altura_var.set(cfg["altura_mm"])
        self.espacamento_linhas_var.set(cfg["espacamento_pt"])
        self.escala_fonte_var.set(cfg["escala_fonte"])

    def _preencher_cred_por_origem(self) -> None:
        if not self._modo_rede_ativo():
            return
        origem = self._valor_listbox(self.lb_origem, "Origem", exibir_erro=False)
        if not origem:
            return
        cred_code = self.origem_para_cred.get(origem, "")
        if not cred_code:
            return
        cred_label = _label_por_codigo_cred(cred_code)
        try:
            indice = CRED_CODES.index(cred_label)
        except ValueError:
            return
        self.lb_cred.selection_clear(0, tk.END)
        self.lb_cred.selection_set(indice)
        self.lb_cred.see(indice)

    @staticmethod
    def _apenas_numeros(valor: str) -> str:
        return re.sub(r"\D", "", valor)

    def _valor_listbox(
        self, listbox: tk.Listbox, campo: str, exibir_erro: bool = True
    ) -> str:
        selecao = listbox.curselection()
        if not selecao:
            if exibir_erro:
                messagebox.showerror(MSG_CAMPO_OBRIGATORIO, f"Selecione um valor para {campo}.")
            return ""
        return listbox.get(selecao[0])

    def _validar_campo_numerico(
        self, campo: str, valor: str, max_digitos: int | None = None
    ) -> bool:
        if not valor:
            messagebox.showerror(MSG_CAMPO_OBRIGATORIO, f"Informe o campo {campo}.")
            return False
        if not valor.isdigit():
            messagebox.showerror(MSG_CAMPO_INVALIDO, f"{campo} aceita apenas numeros.")
            return False
        if max_digitos is not None and len(valor) > max_digitos:
            messagebox.showerror(
                MSG_CAMPO_INVALIDO, f"{campo} aceita somente numeros (max {max_digitos})."
            )
            return False
        return True

    def _validar_volume_total(self, volume_qtd: str) -> int | None:
        if not self._validar_campo_numerico("Volume", volume_qtd, 3):
            return None
        total_volumes = int(volume_qtd)
        if total_volumes <= 0:
            messagebox.showerror(MSG_CAMPO_INVALIDO, "O campo Volume deve ser maior que zero.")
            return None
        return total_volumes

    def _coletar_dados_rede(
        self, origem: str, destino: str, projeto: str, data_emissao: str
    ) -> dict | None:
        tecnologia = self.tecnologia_var.get().strip().upper()
        nota_fiscal = self.nota_fiscal_var.get().strip()
        os_num = self.os_var.get().strip()
        id_fedex = self.id_fedex_var.get().strip()
        volume_qtd = self.volume_qtd_var.get().strip()
        cred_label = self._valor_listbox(self.lb_cred, "Numero CRED")
        if not cred_label:
            return None

        if not tecnologia:
            messagebox.showerror(MSG_CAMPO_OBRIGATORIO, "Informe o campo Tecnologia.")
            return None
        if len(tecnologia) > 3:
            messagebox.showerror(MSG_CAMPO_INVALIDO, "Tecnologia permite no maximo 3 caracteres.")
            return None
        if not re.fullmatch(r"[A-Za-z]{1,3}", tecnologia):
            messagebox.showerror(MSG_CAMPO_INVALIDO, "Tecnologia aceita apenas texto (letras).")
            return None
        if not self._validar_campo_numerico("Nota Fiscal", nota_fiscal, 8):
            return None
        if not self._validar_campo_numerico("OS", os_num, 10):
            return None
        if not self._validar_campo_numerico("ID FEDEX", id_fedex, 10):
            return None

        total_volumes = self._validar_volume_total(volume_qtd)
        if total_volumes is None:
            return None

        vol_total_fmt = str(total_volumes).zfill(3)
        etiquetas = []
        for indice in range(1, total_volumes + 1):
            vol_atual_fmt = str(indice).zfill(3)
            volume_num = f"{vol_atual_fmt}{vol_total_fmt}"
            etiquetas.append(
                {
                    "mode": "REDE",
                    "titulo": "OPERACAO REVERSA",
                    "tecnologia": tecnologia,
                    "origem": origem,
                    "destino": destino,
                    "projeto": projeto,
                    "numero_cred_label": cred_label,
                    "nota_fiscal": nota_fiscal,
                    "data_emissao": data_emissao,
                    "os": os_num,
                    "id_fedex": id_fedex,
                    "volume": f"{vol_atual_fmt}/{vol_total_fmt}",
                    "codigo_barras": f"{nota_fiscal}{os_num}{volume_num}",
                }
            )
        return {
            "mode": "REDE",
            "origem": origem,
            "destino": destino,
            "projeto": projeto,
            "volume_total": total_volumes,
            "etiquetas": etiquetas,
        }

    def _coletar_dados_padrao(
        self, origem: str, destino: str, projeto: str, data_emissao: str
    ) -> dict | None:
        sufixo_romaneio = self.romaneio_sufixo_var.get().strip()
        nr_nf = self.nr_nf_var.get().strip()
        id_fedex = self.id_fedex_var.get().strip()
        volume_qtd = self.volume_qtd_var.get().strip()

        if not sufixo_romaneio:
            messagebox.showerror(MSG_CAMPO_OBRIGATORIO, "Informe os numeros do Romaneio.")
            return None
        if not sufixo_romaneio.isdigit():
            messagebox.showerror(MSG_CAMPO_INVALIDO, "Romaneio aceita apenas numeros.")
            return None
        if not self._validar_campo_numerico("NR NF", nr_nf):
            return None
        if not self._validar_campo_numerico("ID FEDEX", id_fedex, 10):
            return None

        total_volumes = self._validar_volume_total(volume_qtd)
        if total_volumes is None:
            return None

        romaneio = f"{PREFIXOS_ROMANEIO[projeto]}{sufixo_romaneio}"
        base_codigo = self._apenas_numeros(romaneio)
        vol_total_fmt = str(total_volumes).zfill(3)
        id_fedex_data = f"{id_fedex} - {data_emissao}"

        etiquetas = []
        for indice in range(1, total_volumes + 1):
            vol_atual_fmt = str(indice).zfill(3)
            volume_fmt = f"{vol_atual_fmt}/{vol_total_fmt}"
            etiquetas.append(
                {
                    "mode": "PADRAO",
                    "origem": origem,
                    "destino": destino,
                    "projeto": projeto,
                    "romaneio": romaneio,
                    "nr_nf": nr_nf,
                    "id_fedex_data": id_fedex_data,
                    "volume": volume_fmt,
                    "codigo_barras": f"{base_codigo}{vol_atual_fmt}{vol_total_fmt}",
                }
            )

        return {
            "mode": "PADRAO",
            "origem": origem,
            "destino": destino,
            "projeto": projeto,
            "romaneio": romaneio,
            "nr_nf": nr_nf,
            "id_fedex_data": id_fedex_data,
            "volume_total": total_volumes,
            "etiquetas": etiquetas,
        }

    def _coletar_dados(self) -> dict | None:
        origem = self._valor_listbox(self.lb_origem, "Origem")
        destino = self._valor_listbox(self.lb_destino, "Destino")
        projeto = self._valor_listbox(self.lb_projeto, "Projeto")
        if not origem or not destino or not projeto:
            return None

        data_emissao = datetime.now().strftime("%d/%m/%Y")
        if projeto == PROJETO_REDE:
            return self._coletar_dados_rede(origem, destino, projeto, data_emissao)
        return self._coletar_dados_padrao(origem, destino, projeto, data_emissao)

    def _atualizar_preview(self, dados: dict) -> None:
        etiquetas = dados["etiquetas"]
        total = len(etiquetas)
        if dados["mode"] == "REDE":
            e = etiquetas[0]
            limite_preview = 40
            volumes_preview = [item["volume"] for item in etiquetas[:limite_preview]]
            codigos_preview = [
                f"{item['volume']} -> {item['codigo_barras']}" for item in etiquetas[:limite_preview]
            ]
            if total > limite_preview:
                volumes_preview.append(
                    f"... (mostrando {limite_preview} de {total} etiquetas)"
                )
                codigos_preview.append(
                    f"... (mostrando {limite_preview} de {total} etiquetas)"
                )
            texto = (
                f"Titulo: {e['titulo']}\n"
                f"Tecnologia: {e['tecnologia']}\n"
                f"Origem: {e['origem']}\n"
                f"Destino: {e['destino']}\n"
                f"Numero CRED: {e['numero_cred_label']}\n"
                f"Nota Fiscal: {e['nota_fiscal']}\n"
                f"Data Emissao: {e['data_emissao']}\n"
                f"OS: {e['os']}\n"
                f"ID FEDEX: {e['id_fedex']}\n"
                f"Quantidade de etiquetas: {total}\n"
                "Volumes:\n"
                + "\n".join(volumes_preview)
                + "\n\nVolumes / Codigos:\n"
                + "\n".join(codigos_preview)
            )
        else:
            limite_preview = 80
            linhas_preview = [
                f"{item['volume']} -> {item['codigo_barras']}"
                for item in etiquetas[:limite_preview]
            ]
            if total > limite_preview:
                linhas_preview.append(
                    f"... (mostrando {limite_preview} de {total} etiquetas)"
                )

            texto = (
                f"Origem: {dados['origem']}\n"
                f"Destino: {dados['destino']}\n"
                f"Projeto: {dados['projeto']}\n"
                f"Romaneio: {dados['romaneio']}\n"
                f"NR NF: {dados['nr_nf']}\n"
                f"ID FEDEX: {dados['id_fedex_data']}\n"
                f"Quantidade de etiquetas: {total}\n\n"
                "Volumes / Codigos de barras:\n"
                + "\n".join(linhas_preview)
            )
        self.txt_preview.config(state="normal")
        self.txt_preview.delete("1.0", tk.END)
        self.txt_preview.insert("1.0", texto)
        self.txt_preview.config(state="disabled")

    def _atualizar_campo_codigo(self, dados: dict) -> None:
        if dados["mode"] == "REDE":
            etiquetas = dados["etiquetas"]
            primeiro = etiquetas[0]["codigo_barras"]
            if len(etiquetas) == 1:
                self.codigo_barras_var.set(primeiro)
            else:
                self.codigo_barras_var.set(f"{primeiro} (+{len(etiquetas) - 1})")
            return
        etiquetas = dados["etiquetas"]
        primeiro = etiquetas[0]["codigo_barras"]
        if len(etiquetas) == 1:
            self.codigo_barras_var.set(primeiro)
        else:
            self.codigo_barras_var.set(f"{primeiro} (+{len(etiquetas) - 1})")

    def gerar_codigo(self) -> None:
        dados = self._coletar_dados()
        if not dados:
            return
        self._atualizar_campo_codigo(dados)
        self._atualizar_preview(dados)

    def _tamanho_etiqueta_points(self) -> tuple[float, float] | tuple[None, None]:
        largura_txt = self.etiqueta_largura_var.get().strip().replace(",", ".")
        altura_txt = self.etiqueta_altura_var.get().strip().replace(",", ".")
        try:
            largura_mm = float(largura_txt)
            altura_mm = float(altura_txt)
        except ValueError:
            messagebox.showerror(MSG_TAMANHO_INVALIDO, "Informe largura/altura validas em mm.")
            return None, None

        if largura_mm <= 0 or altura_mm <= 0:
            messagebox.showerror(
                MSG_TAMANHO_INVALIDO, "Largura e altura da etiqueta devem ser maiores que zero."
            )
            return None, None
        return largura_mm * MM_TO_POINTS, altura_mm * MM_TO_POINTS

    def _ajustes_layout(self) -> tuple[float, float, float, float] | tuple[None, None, None, None]:
        espacamento_txt = self.espacamento_linhas_var.get().strip().replace(",", ".")
        escala_fonte_txt = self.escala_fonte_var.get().strip().replace(",", ".")
        ajuste_cabecalho_txt = self.ajuste_cabecalho_var.get().strip().replace(",", ".")
        ajuste_rodape_txt = self.ajuste_rodape_var.get().strip().replace(",", ".")
        try:
            espacamento_extra = float(espacamento_txt)
            escala_fonte = float(escala_fonte_txt)
            ajuste_cabecalho = float(ajuste_cabecalho_txt)
            ajuste_rodape = float(ajuste_rodape_txt)
        except ValueError:
            messagebox.showerror(
                MSG_AJUSTE_INVALIDO,
                "Informe valores validos para Espacamento, Escala fonte, Ajuste cabecalho e Ajuste rodape.",
            )
            return None, None, None, None

        if espacamento_extra < 0:
            messagebox.showerror(MSG_AJUSTE_INVALIDO, "Espacamento deve ser maior ou igual a zero.")
            return None, None, None, None
        if escala_fonte <= 0:
            messagebox.showerror(MSG_AJUSTE_INVALIDO, "Escala fonte deve ser maior que zero.")
            return None, None, None, None
        if ajuste_cabecalho < 0:
            messagebox.showerror(
                MSG_AJUSTE_INVALIDO, "Ajuste cabecalho deve ser maior ou igual a zero."
            )
            return None, None, None, None
        if ajuste_rodape < 0:
            messagebox.showerror(MSG_AJUSTE_INVALIDO, "Ajuste rodape deve ser maior ou igual a zero.")
            return None, None, None, None
        return espacamento_extra, escala_fonte, ajuste_cabecalho, ajuste_rodape

    def _layout_paginas_a4(
        self, largura_pt: float, altura_pt: float, exibir_erro: bool = True
    ) -> dict | None:
        pagina_largura, pagina_altura = A4
        margem = 12 * MM_TO_POINTS
        gap = 4 * MM_TO_POINTS
        area_largura = pagina_largura - (2 * margem)
        area_altura = pagina_altura - (2 * margem)

        colunas = int((area_largura + gap) // (largura_pt + gap))
        linhas = int((area_altura + gap) // (altura_pt + gap))

        if colunas < 1 or linhas < 1:
            if exibir_erro:
                messagebox.showerror(
                    MSG_TAMANHO_INVALIDO,
                    "Com esse tamanho de etiqueta nao cabe nenhuma unidade na folha A4.\n"
                    "Reduza largura/altura.",
                )
            return None

        x_inicial = margem
        y_inicial = pagina_altura - margem - altura_pt
        passo_x = largura_pt + gap
        passo_y = altura_pt + gap

        posicoes = []
        for linha in range(linhas):
            for coluna in range(colunas):
                x = x_inicial + (coluna * passo_x)
                y = y_inicial - (linha * passo_y)
                posicoes.append((x, y))

        return {
            "colunas": colunas,
            "linhas": linhas,
            "por_pagina": colunas * linhas,
            "posicoes": posicoes,
        }

    def _desenhar_etiqueta_pdf(
        self,
        c,
        x: float,
        y: float,
        largura_pt: float,
        altura_pt: float,
        dados: dict,
        espacamento_extra: float,
        escala_fonte_usuario: float,
        ajuste_cabecalho: float = 0.0,
    ) -> None:
        draw_template_padrao(
            c=c,
            barcode_module=code128,
            x=x,
            y=y,
            largura_pt=largura_pt,
            altura_pt=altura_pt,
            dados=dados,
            mm_to_points=MM_TO_POINTS,
            app_name=APP_NAME,
            espacamento_extra=espacamento_extra,
            escala_fonte_usuario=escala_fonte_usuario,
            ajuste_cabecalho=ajuste_cabecalho,
        )

    def _desenhar_etiqueta_rede_pdf(
        self,
        c,
        x: float,
        y: float,
        largura_pt: float,
        altura_pt: float,
        dados: dict,
        espacamento_extra: float,
        escala_fonte_usuario: float,
        ajuste_cabecalho: float = 0.0,
        ajuste_rodape: float = 0.0,
    ) -> None:
        draw_template_rede(
            c=c,
            barcode_module=code128,
            x=x,
            y=y,
            largura_pt=largura_pt,
            altura_pt=altura_pt,
            dados=dados,
            mm_to_points=MM_TO_POINTS,
            espacamento_extra=espacamento_extra,
            escala_fonte_usuario=escala_fonte_usuario,
            ajuste_cabecalho=ajuste_cabecalho,
            ajuste_rodape=ajuste_rodape,
        )

    def _gerar_pdf_etiqueta(self, caminho_pdf: Path, dados_lote: dict) -> bool:
        if not REPORTLAB_AVAILABLE:
            messagebox.showerror(
                "Dependencia ausente",
                "Biblioteca reportlab nao encontrada.\n"
                "Instale com: pip install reportlab",
            )
            return False

        largura_pt, altura_pt = self._tamanho_etiqueta_points()
        if not largura_pt or not altura_pt:
            return False
        (
            espacamento_extra,
            escala_fonte_usuario,
            ajuste_cabecalho,
            ajuste_rodape,
        ) = self._ajustes_layout()
        if (
            espacamento_extra is None
            or escala_fonte_usuario is None
            or ajuste_cabecalho is None
            or ajuste_rodape is None
        ):
            return False
        layout = self._layout_paginas_a4(largura_pt, altura_pt, exibir_erro=True)
        if not layout:
            return False

        try:
            caminho_pdf = Path(caminho_pdf)
            caminho_pdf.parent.mkdir(parents=True, exist_ok=True)

            c = canvas.Canvas(str(caminho_pdf), pagesize=A4)
            c.setTitle(APP_NAME)

            for indice, etiqueta in enumerate(dados_lote["etiquetas"]):
                indice_slot = indice % layout["por_pagina"]
                if indice > 0 and indice_slot == 0:
                    c.showPage()

                x, y = layout["posicoes"][indice_slot]
                if etiqueta["mode"] == "REDE":
                    self._desenhar_etiqueta_rede_pdf(
                        c,
                        x,
                        y,
                        largura_pt,
                        altura_pt,
                        etiqueta,
                        espacamento_extra,
                        escala_fonte_usuario,
                        ajuste_cabecalho,
                        ajuste_rodape,
                    )
                else:
                    self._desenhar_etiqueta_pdf(
                        c,
                        x,
                        y,
                        largura_pt,
                        altura_pt,
                        etiqueta,
                        espacamento_extra,
                        escala_fonte_usuario,
                        ajuste_cabecalho,
                    )

            c.showPage()

            c.save()
            return True
        except Exception as erro:
            messagebox.showerror(
                "Falha ao salvar PDF",
                "Nao foi possivel salvar o PDF no caminho selecionado.\n\n"
                f"Caminho: {caminho_pdf}\n"
                f"Detalhe: {erro}",
            )
            return False

    def salvar_pdf(self) -> None:
        dados = self._coletar_dados()
        if not dados:
            return
        self._atualizar_campo_codigo(dados)
        self._atualizar_preview(dados)

        caminho = filedialog.asksaveasfilename(
            title="Salvar etiqueta em PDF",
            defaultextension=".pdf",
            filetypes=[("Arquivos PDF", "*.pdf")],
            initialfile=f"etiqueta_{datetime.now():%Y%m%d_%H%M%S}.pdf",
        )
        if not caminho:
            return

        if self._gerar_pdf_etiqueta(Path(caminho), dados):
            largura_pt, altura_pt = self._tamanho_etiqueta_points()
            layout = None
            if largura_pt and altura_pt:
                layout = self._layout_paginas_a4(largura_pt, altura_pt, exibir_erro=False)
            por_folha = layout["por_pagina"] if layout else 1
            total = len(dados["etiquetas"])
            folhas = (total + por_folha - 1) // por_folha
            messagebox.showinfo(
                "PDF gerado",
                "Etiquetas salvas em:\n"
                f"{caminho}\n\n"
                f"Quantidade: {total}\n"
                f"Etiquetas por folha: {por_folha}\n"
                f"Folhas estimadas: {folhas}",
            )

    def _carregar_impressoras(self) -> None:
        impressoras = [PRINTER_DEFAULT_LABEL]
        if PYWIN32_AVAILABLE:
            try:
                flags = (
                    win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
                )
                for impressora in win32print.EnumPrinters(flags):
                    nome = impressora[2]
                    if nome and nome not in impressoras:
                        impressoras.append(nome)
            except Exception:
                pass

        self.cmb_impressora["values"] = impressoras
        if self.impressora_var.get() not in impressoras:
            self.impressora_var.set(PRINTER_DEFAULT_LABEL)

    def imprimir(self) -> None:
        dados = self._coletar_dados()
        if not dados:
            return
        self._atualizar_campo_codigo(dados)
        self._atualizar_preview(dados)

        arquivo_temp = (
            Path(tempfile.gettempdir())
            / f"etiqueta_coleta_{datetime.now():%Y%m%d_%H%M%S}.pdf"
        )
        if not self._gerar_pdf_etiqueta(arquivo_temp, dados):
            return

        if not PYWIN32_AVAILABLE:
            messagebox.showwarning(
                "Impressao nao disponivel",
                "Biblioteca pywin32 nao encontrada.\n"
                f"O PDF foi gerado em:\n{arquivo_temp}\n\n"
                "Instale com: pip install pywin32",
            )
            return

        impressora = self.impressora_var.get()
        try:
            if impressora == PRINTER_DEFAULT_LABEL:
                win32api.ShellExecute(0, "print", str(arquivo_temp), None, ".", 0)
            else:
                win32api.ShellExecute(
                    0, "printto", str(arquivo_temp), f'"{impressora}"', ".", 0
                )
            largura_pt, altura_pt = self._tamanho_etiqueta_points()
            layout = None
            if largura_pt and altura_pt:
                layout = self._layout_paginas_a4(largura_pt, altura_pt, exibir_erro=False)
            por_folha = layout["por_pagina"] if layout else 1
            total = len(dados["etiquetas"])
            folhas = (total + por_folha - 1) // por_folha
            messagebox.showinfo(
                "Impressao enviada",
                "Etiquetas enviadas para impressao.\n"
                f"Impressora: {impressora}\n"
                f"Quantidade: {total}\n"
                f"Etiquetas por folha: {por_folha}\n"
                f"Folhas estimadas: {folhas}",
            )
        except Exception as erro:
            messagebox.showerror(
                "Falha na impressao",
                "Nao foi possivel enviar para impressao.\n"
                f"Detalhe: {erro}\n\n"
                f"PDF gerado em:\n{arquivo_temp}",
            )


def main() -> None:
    root = tk.Tk()
    EtiquetaColetaApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()

